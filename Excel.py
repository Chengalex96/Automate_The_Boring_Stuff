# Reading Excel Spreadsheets

import openpyxl
import os
os.chdir(r'C:\Users\alex_\.vscode\Python\Automate the Boring Stuff\Automate_The_Boring_Stuff\excelSpreadsheets')

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

sheet = workbook['Sheet1']
print(type(sheet))

print(workbook.sheetnames)

print(sheet['A1'].value)

print(sheet.cell(row = 2, column=2).value)

for i in range(1,8):
    print(i, sheet.cell(row = i, column = 1).value)
