import openpyxl
wb = openpyxl.Workbook() # Creates a workbook

print(wb.sheetnames)

sheet = wb['Sheet']

sheet['A1'] = 42
sheet['A2'] = 'Hello'

import os
os.chdir(r'C:\Users\alex_\.vscode\Python\Automate the Boring Stuff\Automate_The_Boring_Stuff\excelSpreadsheets')

wb.save('Example2.xlsx')

sheet2 = wb.create_sheet() # Adds a new sheet, does not rename it

print(wb.sheetnames)
print(sheet2.title)
sheet2.title = 'My new sheet name'

print(sheet2.title)
print(wb.sheetnames)

wb.save('Example2.xlsx')

wb.create_sheet(index = 0, title = 'My other sheet')
wb.save('Example3.xlsx')